from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import JSONResponse
import os
import re
from pydantic import BaseModel
import google.generativeai as genai
import fitz
from openpyxl import Workbook, load_workbook
from PIL import Image
import json
import base64

# === CONFIGURATION ===
EXCEL_FILE = "Invoice DB.xlsx"
GOOGLE_API_KEY = "AIzaSyD5vU4Po9_9eAR79J61AqNBZIDsHqvB8Ws"  # <-- Replace with env variable in prod

genai.configure(api_key=GOOGLE_API_KEY)
model = genai.GenerativeModel(model_name="gemini-1.5-flash")
app = FastAPI()

class Base64PDFRequest(BaseModel):
    file_base64: str

# === UTIL: Convert PDF bytes to image bytes
def convert_pdf_to_image_bytes(pdf_bytes):
    image_bytes_list = []
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        image_bytes_list.append(pix.tobytes("png"))

    return image_bytes_list

# === UTIL: Call Gemini to extract invoice data
def extract_invoice_data_with_gemini(image_bytes):
    prompt = """
You are an invoice processing assistant. Extract the following fields from this invoice image and return them in JSON:
- Invoice Number
- Name (Vendor or Recipient)
- Email
- Phone Number
- Total Amount
- Date
- Items (brief description of the products/services)

Respond in compact JSON only. Do not explain anything.
"""
    try:
        response = model.generate_content([
            prompt,
            *[{"mime_type": "image/png", "data": img} for img in image_bytes]
        ])
        text = response.text.strip()
        json_str = re.search(r'\{.*\}', text, re.DOTALL).group(0)
        data = json.loads(json_str)
        return data
    except Exception as e:
        print(f"[ERROR] Gemini parse failure: {e}")
        return None

# === Excel update function
def update_excel(data, excel_path=EXCEL_FILE):
    try:
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(['Invoice Number', 'Name', 'Email', 'Phone Number', 'Total Amount', 'Date', 'Items'])
        else:
            wb = load_workbook(excel_path)
            ws = wb["Sheet1"]

        items = data.get('Items', 'N/A')
        if isinstance(items, list):
            items = "; ".join(items)

        ws.append([
            data.get('Invoice Number', 'Unknown'),
            data.get('Name', 'N/A'),
            data.get('Email', 'N/A'),
            data.get('Phone Number', 'N/A'),
            data.get('Total Amount', '0.00'),
            data.get('Date', 'N/A'),
            items
        ])
        wb.save(excel_path)
        return True
    except Exception as e:
        print(f"[ERROR] Excel write failed: {e}")
        return False

async def process_invoice(file):
    try:
        # Decode base64 PDF
        pdf_bytes = base64.b64decode(file.file_base64)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid base64 input: {e}")

    image_bytes_list = convert_pdf_to_image_bytes(pdf_bytes)
    data = extract_invoice_data_with_gemini(image_bytes_list)

    if not data:
        raise HTTPException(status_code=500, detail="Failed to extract invoice data.")

    # success = update_excel(data)
    # if not success:
    #     raise HTTPException(status_code=500, detail="Failed to update Excel.")

    return JSONResponse(content=data)
